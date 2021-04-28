from tkinter import *
import tkinter as tk
from openpyxl import load_workbook
from tkinter import ttk,filedialog,messagebox
import os
from PIL import Image,ImageTk
from openpyxl.utils.cell import get_column_letter
import openpyxl

class Masala:
    def __init__(self,root):
        self.root=root
        self.root.overrideredirect(True)
        self.root.geometry("{0}x{1}+0+0".format(self.root.winfo_screenwidth(), self.root.winfo_screenheight()))

        self.box_item = StringVar()
        self.no = StringVar()
        self.File_name = StringVar()
        self.var_Rec_No = StringVar()
        self.var_cust_name = StringVar()
        self.var_email_add = StringVar()
        self.var_res_add = StringVar()
        self.var_city_1 = StringVar()
        self.var_state_1 = StringVar()
        self.var_zip = StringVar()
        self.var_ph_no_1 = StringVar()
        self.var_country_1 = StringVar()
        self.var_sex_1 = StringVar()
        self.var_d_birth = StringVar()
        self.var_height = StringVar()
        self.var_weidth = StringVar()
        self.var_Blood_gp = StringVar()
        self.var_billing_name = StringVar()
        self.var_shipper_name = StringVar()
        self.var_city_2 = StringVar()
        self.var_state_2 = StringVar()
        self.var_zip_2 = StringVar()
        self.var_country_2 = StringVar()
        self.var_ph_no_2 = StringVar()
        self.var_alcoholic = StringVar()
        self.var_smoker = StringVar()
        self.var_part_sung = StringVar()
        self.var_diabetic = StringVar()
        self.var_allergised = StringVar()
        self.var_policy_no = StringVar()
        self.var_D_b_lifeassure = StringVar()
        self.var_p_inst = StringVar()
        self.var_name_p_holder = StringVar()
        self.var_stm_name = StringVar()
        self.var_smt_code = StringVar()
        self.var_dob = StringVar()
        self.var_sex_2 = StringVar()
        self.var_crd_name = StringVar()
        self.var_medicine = StringVar()
        self.var_Dosage = StringVar()
        self.var_Tablets = StringVar()
        self.var_pill_rate = StringVar()
        self.var_cost = StringVar()
        self.var_shipimg_cost = StringVar()
        self.var_Total_amount = StringVar()
        self.var_Remark = StringVar()

        self.save_icon = PhotoImage(file=r".\res\save.png")
        self.image_icon = PhotoImage(file=r".\res\im.png")
        self.show_icon = PhotoImage(file=r".\res\sd.png")
        self.bp_icon = PhotoImage(file=r".\res\db.png")
        self.re_icon = PhotoImage(file=r".\res\re3.png")
        self.ext_icon = PhotoImage(file=r".\res\sw.png")

        # ................................Fist Fram btn......................
        firs_fram = Frame(self.root, relief=RIDGE, bg="#053075", bd=4)
        firs_fram.place(x=10, y=5, width=180, height=450)

        Select_Images = Button(firs_fram, text="Select Images   ", width=150, height=50, bd=2,
                               font=("time new roman", 12, "bold"), image=self.image_icon,
                               compound=LEFT, command=self.image_file).grid(row=0, padx=10, pady=8, sticky="w")
        show_data = Button(firs_fram, text="     Show Data", width=150, height=50, bd=2,
                           font=("time new roman", 12, "bold"), image=self.show_icon, compound=LEFT,
                           command=self.new_windows
                           ).grid(
            row=1, padx=10, pady=8, sticky="w")
        btn_backup = Button(firs_fram, text="     Backup     ", width=150, height=50, bd=2,
                            font=("time new roman", 12, "bold"), image=self.bp_icon, compound=LEFT
                            ).grid(
            row=2, padx=10, pady=8, sticky="w")
        btn_resate = Button(firs_fram, text="   Restore      ", width=150, height=50, bd=2,
                            font=("time new roman", 12, "bold"), image=self.re_icon,
                            compound=LEFT, command=self.rest).grid(
            row=3, padx=10, pady=8, sticky="w")
        btn_exit = Button(firs_fram, text="     Exit             ", width=150, height=50, bd=2,
                          font=("time new roman", 12, "bold"), image=self.ext_icon,
                          compound=LEFT, command=self.ext).grid(
            row=4, padx=10, pady=8, sticky="w")

        # ................................secand Fram......................

        secand_fram = Frame(self.root, relief=RIDGE, bg="#053075", bd=4)
        secand_fram.place(x=190, y=5, width=1165, height=450)
        self.File_name = StringVar()

        lbl_Image_Name = Label(secand_fram, text="Image Name", font=("time new roman",9, "bold"), bg="#053075",
                               fg="white")
        lbl_Image_Name.grid(row=0, column=0, pady=5, padx=3, sticky="w")

        txt_Image_Name = Entry(secand_fram, font=("time new roman", 12), bd=1,
                               relief=GROOVE, textvariable=self.File_name)
        txt_Image_Name.grid(row=0, column=1, pady=5, padx=3, sticky="w")

        lbl_Record_No = Label(secand_fram, text="Record No.", font=("time new roman", 9, "bold"), bg="#053075",
                              fg="white")
        lbl_Record_No.grid(row=0, column=2, pady=5, padx=3, sticky="w")

        txt_Record_No = Entry(secand_fram, font=("time new roman", 12), bd=2,
                              relief=GROOVE, textvariable=self.var_Rec_No)
        txt_Record_No.grid(row=0, column=3, pady=5, padx=3, sticky="w")

        lbl_Customer_Name = Label(secand_fram, text="Customer Name", font=("time new roman", 9, "bold"), bg="#053075",
                                  fg="white")
        lbl_Customer_Name.grid(row=0, column=4, pady=5, padx=3, sticky="w")

        txt_Customer_Name = Entry(secand_fram, font=("time new roman", 12), bd=2,
                                  relief=GROOVE, textvariable=self.var_cust_name)
        txt_Customer_Name.grid(row=0, column=5, pady=5, padx=3, sticky="w")

        lbl_Email_Address = Label(secand_fram, text="E-mail Address", font=("time new roman", 9, "bold"), bg="#053075",
                                  fg="white")
        lbl_Email_Address.grid(row=0, column=6, pady=5, padx=3, sticky="w")

        txt_Email_Address = Entry(secand_fram, font=("time new roman", 12), bd=2,
                                  relief=GROOVE, textvariable=self.var_email_add)
        txt_Email_Address.grid(row=0, column=7, pady=5, padx=3, sticky="w")
        lbl_Res_Address = Label(secand_fram, text="Res Address", font=("time new roman", 9, "bold"), bg="#053075",
                                fg="white")
        lbl_Res_Address.grid(row=1, column=0, pady=5, padx=3, sticky="w")

        txt_Res_Address = Entry(secand_fram, font=("time new roman", 12), bd=2,
                                relief=GROOVE, textvariable=self.var_res_add)
        txt_Res_Address.grid(row=1, column=1, pady=5, padx=3, sticky="w")

        lbl_City_1 = Label(secand_fram, text="City_1", font=("time new roman", 9, "bold"), bg="#053075", fg="white")
        lbl_City_1.grid(row=1, column=2, pady=5, padx=3, sticky="w")

        txt_City_1 = Entry(secand_fram, font=("time new roman", 12), bd=2,
                           relief=GROOVE, textvariable=self.var_city_1)
        txt_City_1.grid(row=1, column=3, pady=5, padx=3, sticky="w")

        lbl_State_1 = Label(secand_fram, text="State_1", font=("time new roman", 9, "bold"), bg="#053075", fg="white")
        lbl_State_1.grid(row=1, column=4, pady=5, padx=3, sticky="w")

        txt_State_1 = Entry(secand_fram, font=("time new roman", 12), bd=2,
                            relief=GROOVE, textvariable=self.var_state_1)
        txt_State_1.grid(row=1, column=5, pady=5, padx=3, sticky="w")

        lbl_Zip = Label(secand_fram, text="Zip", font=("time new roman", 9, "bold"), bg="#053075", fg="white")
        lbl_Zip.grid(row=1, column=6, pady=5, padx=3, sticky="w")

        txt_Zip = Entry(secand_fram, font=("time new roman", 12), bd=2,
                        relief=GROOVE, textvariable=self.var_zip)
        txt_Zip.grid(row=1, column=7, pady=5, padx=3, sticky="w")

        lbl_Ph_No_1 = Label(secand_fram, text="Ph_No_1", font=("time new roman", 9, "bold"), bg="#053075", fg="white")
        lbl_Ph_No_1.grid(row=2, column=0, pady=5, padx=3, sticky="w")

        txt_Ph_No_1 = Entry(secand_fram, font=("time new roman", 12), bd=2,
                            relief=GROOVE, textvariable=self.var_ph_no_1)
        txt_Ph_No_1.grid(row=2, column=1, pady=5, padx=3, sticky="w")

        lbl_Country_1 = Label(secand_fram, text="Country_1", font=("time new roman", 9, "bold"), bg="#053075",
                              fg="white")
        lbl_Country_1.grid(row=2, column=2, pady=5, padx=3, sticky="w")

        txt_Country_1 = Entry(secand_fram, font=("time new roman", 12), bd=2,
                              relief=GROOVE, textvariable=self.var_country_1)
        txt_Country_1.grid(row=2, column=3, pady=5, padx=3, sticky="w")

        lbl_Sex_1 = Label(secand_fram, text="Sex_1", font=("time new roman", 9, "bold"), bg="#053075", fg="white")
        lbl_Sex_1.grid(row=2, column=4, pady=5, padx=3, sticky="w")

        txt_Sex_1 = Entry(secand_fram, font=("time new roman", 12), bd=2,
                          relief=GROOVE, textvariable=self.var_sex_1)
        txt_Sex_1.grid(row=2, column=5, pady=5, padx=3, sticky="w")

        lbl_D_Birth = Label(secand_fram, text="D_Birth", font=("time new roman", 9, "bold"), bg="#053075", fg="white")
        lbl_D_Birth.grid(row=2, column=6, pady=5, padx=3, sticky="w")

        txt_D_Birth = Entry(secand_fram, font=("time new roman", 12), bd=2,
                            relief=GROOVE, textvariable=self.var_d_birth)
        txt_D_Birth.grid(row=2, column=7, pady=5, padx=3, sticky="w")

        lbl_Height = Label(secand_fram, text="Height", font=("time new roman", 9, "bold"), bg="#053075", fg="white")
        lbl_Height.grid(row=3, column=0, pady=5, padx=3, sticky="w")

        txt_Height = Entry(secand_fram, font=("time new roman", 12), bd=2,
                           relief=GROOVE, textvariable=self.var_height)
        txt_Height.grid(row=3, column=1, pady=5, padx=3, sticky="w")

        lbl_Weigth = Label(secand_fram, text="Weigth", font=("time new roman", 9, "bold"), bg="#053075", fg="white")
        lbl_Weigth.grid(row=3, column=2, pady=5, padx=3, sticky="w")

        txt_Weigth = Entry(secand_fram, font=("time new roman", 12), bd=2,
                           relief=GROOVE, textvariable=self.var_weidth)
        txt_Weigth.grid(row=3, column=3, pady=5, padx=3, sticky="w")

        lbl_Blood_Group = Label(secand_fram, text="Blood Group", font=("time new roman", 9, "bold"), bg="#053075",
                                fg="white")

        lbl_Blood_Group.grid(row=3, column=4, pady=5, padx=3, sticky="w")

        txt_Blood_Group = Entry(secand_fram, font=("time new roman", 12), bd=2,

                                relief=GROOVE, textvariable=self.var_Blood_gp)
        txt_Blood_Group.grid(row=3, column=5, pady=5, padx=3, sticky="w")

        lbl_Billing_Name = Label(secand_fram, text="Billing Name", font=("time new roman", 9, "bold"), bg="#053075",
                                 fg="white")
        lbl_Billing_Name.grid(row=3, column=6, pady=5, padx=3, sticky="w")

        txt_Billing_Name = Entry(secand_fram, font=("time new roman", 12), bd=2,
                                 relief=GROOVE, textvariable=self.var_billing_name)
        txt_Billing_Name.grid(row=3, column=7, pady=5, padx=3, sticky="w")

        lbl_Shipper_Name = Label(secand_fram, text="Shipper Name", font=("time new roman", 9, "bold"), bg="#053075",
                                 fg="white")
        lbl_Shipper_Name.grid(row=4, column=0, pady=5, padx=3, sticky="w")

        txt_Shipper_Name = Entry(secand_fram, font=("time new roman", 12), bd=2,
                                 relief=GROOVE, textvariable=self.var_shipper_name)
        txt_Shipper_Name.grid(row=4, column=1, pady=5, padx=3, sticky="w")

        lbl_City_2 = Label(secand_fram, text="City_2", font=("time new roman", 9, "bold"), bg="#053075", fg="white")
        lbl_City_2.grid(row=4, column=2, pady=5, padx=3, sticky="w")

        txt_City_2 = Entry(secand_fram, font=("time new roman", 12), bd=2,
                           relief=GROOVE, textvariable=self.var_city_2)
        txt_City_2.grid(row=4, column=3, pady=5, padx=3, sticky="w")

        lbl_State_2 = Label(secand_fram, text="State_2", font=("time new roman", 9, "bold"), bg="#053075", fg="white")
        lbl_State_2.grid(row=4, column=4, pady=5, padx=3, sticky="w")

        txt_State_2 = Entry(secand_fram, font=("time new roman", 12), bd=2,
                            relief=GROOVE, textvariable=self.var_state_2)
        txt_State_2.grid(row=4, column=5, pady=5, padx=3, sticky="w")

        lbl_Zip_2 = Label(secand_fram, text="Zip_2", font=("time new roman", 9, "bold"), bg="#053075", fg="white")
        lbl_Zip_2.grid(row=4, column=6, pady=5, padx=3, sticky="w")

        txt_Zip_2 = Entry(secand_fram, font=("time new roman", 12), bd=2,
                          relief=GROOVE, textvariable=self.var_zip_2)
        txt_Zip_2.grid(row=4, column=7, pady=5, padx=3, sticky="w")

        lbl_Country_2 = Label(secand_fram, text="Country_2", font=("time new roman", 9, "bold"), bg="#053075",
                              fg="white")
        lbl_Country_2.grid(row=5, column=0, pady=5, padx=3, sticky="w")

        txt_Country_2 = Entry(secand_fram, font=("time new roman", 12), bd=2,
                              relief=GROOVE, textvariable=self.var_country_2)
        txt_Country_2.grid(row=5, column=1, pady=5, padx=3, sticky="w")

        lbl_Ph_No_2 = Label(secand_fram, text="Ph_No_2", font=("time new roman", 9, "bold"), bg="#053075", fg="white")
        lbl_Ph_No_2.grid(row=5, column=2, pady=5, padx=3, sticky="w")

        txt_Ph_No_2 = Entry(secand_fram, font=("time new roman", 12), bd=2,
                            relief=GROOVE, textvariable=self.var_ph_no_2)
        txt_Ph_No_2.grid(row=5, column=3, pady=5, padx=3, sticky="w")

        lbl_Alcoholic = Label(secand_fram, text="Alcoholic", font=("time new roman", 9, "bold"), bg="#053075",
                              fg="white")
        lbl_Alcoholic.grid(row=5, column=4, pady=5, padx=3, sticky="w")

        txt_Alcoholic = Entry(secand_fram, font=("time new roman", 12), bd=2,
                              relief=GROOVE, textvariable=self.var_alcoholic)
        txt_Alcoholic.grid(row=5, column=5, pady=5, padx=3, sticky="w")

        lbl_Smoker = Label(secand_fram, text="Smoker", font=("time new roman", 9, "bold"), bg="#053075", fg="white")
        lbl_Smoker.grid(row=5, column=6, pady=5, padx=3, sticky="w")

        txt_Smoker = Entry(secand_fram, font=("time new roman", 12), bd=2,
                           relief=GROOVE, textvariable=self.var_smoker)
        txt_Smoker.grid(row=5, column=7, pady=5, padx=3, sticky="w")

        lbl_Part_sung = Label(secand_fram, text="Part_sung", font=("time new roman", 9, "bold"), bg="#053075",
                              fg="white")
        lbl_Part_sung.grid(row=6, column=0, pady=5, padx=5, sticky="w")

        txt_Part_sung = Entry(secand_fram, font=("time new roman", 12), bd=2,
                              relief=GROOVE, textvariable=self.var_part_sung)
        txt_Part_sung.grid(row=6, column=1, pady=5, padx=3, sticky="w")

        lbl_Diabetic = Label(secand_fram, text="Diabetic", font=("time new roman", 9, "bold"), bg="#053075",
                             fg="white")
        lbl_Diabetic.grid(row=6, column=2, pady=5, padx=3, sticky="w")

        txt_Diabetic = Entry(secand_fram, font=("time new roman", 12), bd=2,
                             relief=GROOVE, textvariable=self.var_diabetic)
        txt_Diabetic.grid(row=6, column=3, pady=5, padx=3, sticky="w")

        lbl_Allergised = Label(secand_fram, text="Allergised", font=("time new roman", 9, "bold"), bg="#053075",
                               fg="white")
        lbl_Allergised.grid(row=6, column=4, pady=5, padx=3, sticky="w")

        txt_Allergised = Entry(secand_fram, font=("time new roman", 12), bd=2,
                               relief=GROOVE, textvariable=self.var_allergised)
        txt_Allergised.grid(row=6, column=5, pady=5, padx=3, sticky="w")

        lbl_Policy_no = Label(secand_fram, text="Policy_no", font=("time new roman", 9, "bold"), bg="#053075",
                              fg="white")
        lbl_Policy_no.grid(row=6, column=6, pady=5, padx=3, sticky="w")

        txt_Policy_no = Entry(secand_fram, font=("time new roman", 12), bd=2,
                              relief=GROOVE, textvariable=self.var_policy_no)
        txt_Policy_no.grid(row=6, column=7, pady=5, padx=3, sticky="w")

        lbl_D_B_LifeAssure = Label(secand_fram, text="D-B_LifeAssure", font=("time new roman", 9, "bold"),
                                   bg="#053075", fg="white")
        lbl_D_B_LifeAssure.grid(row=7, column=0, pady=5, padx=3, sticky="w")

        txt_D_B_LifeAssure = Entry(secand_fram, font=("time new roman", 12), bd=2,
                                   relief=GROOVE, textvariable=self.var_D_b_lifeassure)
        txt_D_B_LifeAssure.grid(row=7, column=1, pady=5, padx=3, sticky="w")

        lbl_p_Inst = Label(secand_fram, text="p_Inst", font=("time new roman", 9, "bold"), bg="#053075", fg="white")
        lbl_p_Inst.grid(row=7, column=2, pady=5, padx=3, sticky="w")

        txt_p_Inst = Entry(secand_fram, font=("time new roman", 12), bd=2,
                           relief=GROOVE, textvariable=self.var_p_inst)
        txt_p_Inst.grid(row=7, column=3, pady=5, padx=3, sticky="w")

        lbl_Name_p_Holder = Label(secand_fram, text="Name_p_Holder", font=("time new roman", 9, "bold"), bg="#053075",
                                  fg="white")
        lbl_Name_p_Holder.grid(row=7, column=4, pady=5, padx=3, sticky="w")

        txt_Name_p_Holder = Entry(secand_fram, font=("time new roman", 12), bd=2,
                                  relief=GROOVE, textvariable=self.var_name_p_holder)
        txt_Name_p_Holder.grid(row=7, column=5, pady=5, padx=3, sticky="w")

        lbl_STM_Name = Label(secand_fram, text="STM Name", font=("time new roman", 9, "bold"), bg="#053075",
                             fg="white")
        lbl_STM_Name.grid(row=7, column=6, pady=5, padx=3, sticky="w")

        txt_STM_Name = Entry(secand_fram, font=("time new roman", 12), bd=2,
                             relief=GROOVE, textvariable=self.var_stm_name)
        txt_STM_Name.grid(row=7, column=7, pady=5, padx=3, sticky="w")

        lbl_SMT_Code = Label(secand_fram, text="SMT Code", font=("time new roman", 9, "bold"), bg="#053075",
                             fg="white")
        lbl_SMT_Code.grid(row=8, column=0, pady=5, padx=3, sticky="w")

        txt_SMT_Code = Entry(secand_fram, font=("time new roman", 12), bd=2,
                             relief=GROOVE, textvariable=self.var_smt_code)
        txt_SMT_Code.grid(row=8, column=1, pady=5, padx=3, sticky="w")

        lbl_DOB = Label(secand_fram, text="DOB", font=("time new roman", 9), bg="#053075", fg="white")
        lbl_DOB.grid(row=8, column=2, pady=5, padx=3, sticky="w")

        txt_DOB = Entry(secand_fram, font=("time new roman", 12), bd=2,
                        relief=GROOVE, textvariable=self.var_dob)
        txt_DOB.grid(row=8, column=3, pady=5, padx=3, sticky="w")

        lbl_Sex_2 = Label(secand_fram, text="Sex_2", font=("time new roman", 9, "bold"), bg="#053075", fg="white")
        lbl_Sex_2.grid(row=8, column=4, pady=5, padx=3, sticky="w")

        txt_Sex_2 = Entry(secand_fram, font=("time new roman", 12), bd=2,
                          relief=GROOVE, textvariable=self.var_sex_2)
        txt_Sex_2.grid(row=8, column=5, pady=5, padx=3, sticky="w")

        lbl_Card_Name = Label(secand_fram, text="Card Name", font=("time new roman", 9, "bold"), bg="#053075",
                              fg="white")
        lbl_Card_Name.grid(row=8, column=6, pady=5, padx=3, sticky="w")

        txt_Card_Name = Entry(secand_fram, font=("time new roman", 12), bd=2,
                              relief=GROOVE, textvariable=self.var_crd_name)
        txt_Card_Name.grid(row=8, column=7, pady=5, padx=3, sticky="w")

        lbl_Medicine = Label(secand_fram, text="Medicine", font=("time new roman", 9, "bold"), bg="#053075",
                             fg="white")
        lbl_Medicine.grid(row=9, column=0, pady=5, padx=3, sticky="w")

        txt_Medicine = Entry(secand_fram, font=("time new roman", 12), bd=2,
                             relief=GROOVE, textvariable=self.var_medicine)
        txt_Medicine.grid(row=9, column=1, pady=5, padx=3, sticky="w")

        lbl_Dosage = Label(secand_fram, text="Dosage", font=("time new roman", 9, "bold"), bg="#053075", fg="white")
        lbl_Dosage.grid(row=9, column=2, pady=5, padx=3, sticky="w")

        txt_Dosage = Entry(secand_fram, font=("time new roman", 12), bd=2,
                           relief=GROOVE, textvariable=self.var_Dosage)
        txt_Dosage.grid(row=9, column=3, pady=5, padx=3, sticky="w")

        lbl_Tablets = Label(secand_fram, text="Tablets", font=("time new roman", 9, "bold"), bg="#053075", fg="white")
        lbl_Tablets.grid(row=9, column=4, pady=5, padx=3, sticky="w")

        txt_Tablets = Entry(secand_fram, font=("time new roman", 12), bd=2,
                            relief=GROOVE, textvariable=self.var_Tablets)
        txt_Tablets.grid(row=9, column=5, pady=5, padx=3, sticky="w")

        lbl_Pill_Rate = Label(secand_fram, text="Pill Rate", font=("time new roman", 9, "bold"), bg="#053075",
                              fg="white")
        lbl_Pill_Rate.grid(row=9, column=6, pady=5, padx=3, sticky="w")

        txt_Pill_Rate = Entry(secand_fram, font=("time new roman", 12), bd=2,
                              relief=GROOVE, textvariable=self.var_pill_rate)
        txt_Pill_Rate.grid(row=9, column=7, pady=5, padx=3, sticky="w")

        lbl_Cost = Label(secand_fram, text="Cost", font=("time new roman", 9, "bold"), bg="#053075", fg="white")
        lbl_Cost.grid(row=10, column=0, pady=5, padx=3, sticky="w")

        txt_Cost = Entry(secand_fram, font=("time new roman", 12), bd=2,
                         relief=GROOVE, textvariable=self.var_cost)
        txt_Cost.grid(row=10, column=1, pady=5, padx=3, sticky="w")

        lbl_Shiping_Cost = Label(secand_fram, text="Shiping Cost", font=("time new roman", 9, "bold"), bg="#053075",
                                 fg="white")
        lbl_Shiping_Cost.grid(row=10, column=2, pady=5, padx=3, sticky="w")

        txt_Shiping_Cost = Entry(secand_fram, font=("time new roman", 12), bd=2,
                                 relief=GROOVE, textvariable=self.var_shipimg_cost)
        txt_Shiping_Cost.grid(row=10, column=3, pady=5, padx=3, sticky="w")

        lbl_Total_Amount = Label(secand_fram, text="Total Amount", font=("time new roman", 9, "bold"), bg="#053075",
                                 fg="white")
        lbl_Total_Amount.grid(row=10, column=4, pady=5, padx=3, sticky="w")

        txt_Total_Amount = Entry(secand_fram, font=("time new roman", 12), bd=2,
                                 relief=GROOVE, textvariable=self.var_Total_amount)
        txt_Total_Amount.grid(row=10, column=5, pady=5, padx=3, sticky="w")

        lbl_Remarks = Label(secand_fram, text="Remarks.", font=("time new roman", 10, "bold"), bg="#053075", fg="white")
        lbl_Remarks.grid(row=10, column=6, pady=5, padx=3, sticky="w")

        txt_Remarks = Entry(secand_fram, font=("time new roman", 12), bd=2,
                            relief=GROOVE, textvariable=self.var_Remark)
        txt_Remarks.grid(row=10, column=7, pady=5, padx=3, sticky="w")

        btn_save = Button(secand_fram, text="Save Data", width=175, height=40, bd=2,
                          font=("time new roman", 12, "bold"), image=self.save_icon,
                          compound=LEFT, command=self.save_data).grid(row=11, column=7, padx=8, pady=6, sticky="w")

        # .....................third.........................
        third_fram = Frame(self.root, relief=RIDGE, bg="#053075", bd=4)
        third_fram.place(x=10, y=455, width=1345, height=250)

        self.canv = Canvas(third_fram, relief=SUNKEN)
        self.canv.config(width=400, height=200)
        self.canv.config(highlightthickness=7)

        sbarV = Scrollbar(third_fram, orient=VERTICAL)
        sbarH = Scrollbar(third_fram, orient=HORIZONTAL)

        sbarV.config(command=self.canv.yview)
        sbarH.config(command=self.canv.xview)

        self.canv.config(yscrollcommand=sbarV.set)
        self.canv.config(xscrollcommand=sbarH.set)

        sbarV.pack(side=RIGHT, fill=Y)
        sbarH.pack(side=BOTTOM, fill=X)

        self.canv.pack(side=LEFT, expand=True, fill=BOTH)
        self.box_list()
        # ......................................................End front design...................................................................

        # ........................................................Backent Start........................................................................

    def save_data(self):
        workbook_name = '.\res\Rishu.xlsx'
        wb = load_workbook(workbook_name)
        page = wb.active
        self.m = page.max_row
        for self.i in range(1, self.m + 1):
            line = page.cell(column=2, row=self.i)
            self.no.set(str(self.m + 1))

        try:
            ad = self.var_res_add.get()
            if ad != "":
                data = [
                    self.no.get(),
                    self.var_Rec_No.get(),
                    self.var_cust_name.get(),
                    self.var_email_add.get(),
                    self.var_res_add.get(),
                    self.var_city_1.get(),
                    self.var_state_1.get(),
                    self.var_zip.get(),
                    self.var_ph_no_1.get(),
                    self.var_country_1.get(),
                    self.var_sex_1.get(),
                    self.var_d_birth.get(),
                    self.var_height.get(),
                    self.var_weidth.get(),
                    self.var_Blood_gp.get(),
                    self.var_billing_name.get(),
                    self.var_shipper_name.get(),
                    self.var_city_2.get(),
                    self.var_state_2.get(),
                    self.var_zip_2.get(),
                    self.var_country_2.get(),
                    self.var_ph_no_2.get(),
                    self.var_alcoholic.get(),
                    self.var_smoker.get(),
                    self.var_part_sung.get(),
                    self.var_diabetic.get(),
                    self.var_allergised.get(),
                    self.var_policy_no.get(),
                    self.var_D_b_lifeassure.get(),
                    self.var_p_inst.get(),
                    self.var_name_p_holder.get(),
                    self.var_stm_name.get(),
                    self.var_smt_code.get(),
                    self.var_dob.get(),
                    self.var_sex_2.get(),
                    self.var_crd_name.get(),
                    self.var_medicine.get(),
                    self.var_Dosage.get(),
                    self.var_Tablets.get(),
                    self.var_pill_rate.get(),
                    self.var_cost.get(),
                    self.var_shipimg_cost.get(),
                    self.var_Total_amount.get(),
                    self.var_Remark.get()

                ]

                p = []

                workbook_name = '.\res\Rishu.xlsx'
                wb = load_workbook(workbook_name)
                page = wb.active
                m = page.max_row
                for i in range(2, m + 1):
                    line = page.cell(column=2, row=i)
                    mh = line.value
                    p.append(mh)
                # print(p)
                if self.var_Rec_No.get() not in p:
                    new_companies = [data]
                    for info in new_companies:
                        page.append(info)

                    wb.save(filename=workbook_name)
                    self.show_file()
                    self.clear()
                    self.box_list()
                    p = []
                else:
                    self.update()


                    p = []
            else:
                messagebox.showwarning("emty", "Please fill all entry")
        except:
            messagebox.showwarning("open", "Please close excal file")

    def clear(self):
        self.no.set("")
        self.var_Rec_No.set("")
        self.var_cust_name.set("")
        self.var_email_add.set("")
        self.var_res_add.set("")
        self.var_city_1.set("")
        self.var_state_1.set("")
        self.var_zip.set("")
        self.var_ph_no_1.set("")
        self.var_country_1.set("")
        self.var_sex_1.set("")
        self.var_d_birth.set("")
        self.var_height.set("")
        self.var_weidth.set("")
        self.var_Blood_gp.set("")
        self.var_billing_name.set("")
        self.var_shipper_name.set("")
        self.var_city_2.set("")
        self.var_state_2.set("")
        self.var_zip_2.set("")
        self.var_country_2.set("")
        self.var_ph_no_2.set("")
        self.var_alcoholic.set("")
        self.var_smoker.set("")
        self.var_part_sung.set("")
        self.var_diabetic.set("")
        self.var_allergised.set("")
        self.var_policy_no.set("")
        self.var_D_b_lifeassure.set("")
        self.var_p_inst.set("")
        self.var_name_p_holder.set("")
        self.var_stm_name.set("")
        self.var_smt_code.set("")
        self.var_dob.set("")
        self.var_sex_2.set("")
        self.var_crd_name.set("")
        self.var_medicine.set("")
        self.var_Dosage.set("")
        self.var_Tablets.set("")
        self.var_pill_rate.set("")
        self.var_cost.set("")
        self.var_shipimg_cost.set("")
        self.var_Total_amount.set("")
        self.var_Remark.set("")

    def image_file(self):
        file_path = filedialog.askopenfilename(filetype=(("jpg file", "*.jpg"), ("All", "*.*")))
        b = os.path.basename(file_path)
        self.File_name.set(b)

        self.im = Image.open(file_path)
        width, height = self.im.size
        self.canv.config(scrollregion=(0, 0, width, height))
        self.im2 = ImageTk.PhotoImage(self.im.resize((1320, 1200)))
        self.imgtag = self.canv.create_image(0, 0, anchor="nw", image=self.im2)

    def new_windows(self):
        self.top = Toplevel()
        self.top.geometry("1130x500+200+20")
        self.top.title("Show Data")

        # first frame...........................................!
        self.firs_n_frm = Frame(self.top, bg="red", relief=RIDGE, bd=2).place(x=15, y=5, width=1100, height=50)
        combo_s_snd = ttk.Combobox(self.top, width=3, font=("time new roman", 10, "bold"), state="readonly",
                                   textvariable=self.box_item
                                   )
        combo_s_snd['value'] = (self.res)
        # combo_s_snd.current(0)
        combo_s_snd.place(x=400, y=10, width=300, height=40)
        p_btn = Button(self.top, text='Transfer Text', command=self.serch_itm, bg='#5603fc',
                       font=("time new roman", 10, "bold")).place(x=750, y=10, width=100, height=40)

        # secend frame.............................................!
        scnd_frm = Frame(self.top, relief=RIDGE, bd=2)
        scnd_frm.place(x=15, y=60, width=1100, height=400)

        scroll_x = Scrollbar(scnd_frm, orient=HORIZONTAL)
        scroll_y = Scrollbar(scnd_frm, orient=VERTICAL)
        scroll_x.pack(side=BOTTOM, fill=X)
        scroll_y.pack(side=RIGHT, fill=Y)
        self.Data_table = ttk.Treeview(scnd_frm, columns=("NO.",
                                                          "RECORD_NO", "CUSTOMER_NAME", "EMAIL_ADDRESS", "RES_ADDRESS",
                                                          "CITY_1", "STATE_1",
                                                          "ZIP_1", "PH_NO1", "COUNTRY_1", "SEX_1", "D_BIRTH_11",
                                                          "HEIGHT", "WEIGHT", "BLOOD_GROUP", "BILLING_NAME",
                                                          "SHIPPER_NAME", "CITY_2", "STATE_2", "ZIP_2", "COUNTRY_2",
                                                          "PH_NO2", "ALCOHOLIC", "SMOKER",
                                                          "PAST_SURG", "DIABETIC", "ALLERGISED", "POLICY_NO",
                                                          "D_B_LIFE_ASSURE_28", "P_INST", "NAME_P_HOLDER",
                                                          "STM_NAME", "STM_CODE", "DOB_33", "SEX_2", "CARD_NAME",
                                                          "MEDICINE", "DOSAGE", "TABLETS", "PILL_RATE",
                                                          "COST", "SHIPPING_COST", "TOTAL_AMT", "REMARK"),
                                       xscrollcommand=scroll_x.set, yscrollcommand=scroll_y.set)
        self.Data_table.pack(fill=BOTH, expand=1)

        scroll_x.config(command=self.Data_table.xview)
        scroll_y.config(command=self.Data_table.yview)
        self.Data_table.heading("NO.", text="NO.")
        self.Data_table.heading("RECORD_NO", text="RECORD_NO")
        self.Data_table.heading("CUSTOMER_NAME", text="CUSTOMER_NAME")
        self.Data_table.heading("EMAIL_ADDRESS", text="Email")
        self.Data_table.heading("RES_ADDRESS", text="RES_ADDRESS")
        self.Data_table.heading("CITY_1", text="CITY_1")
        self.Data_table.heading("STATE_1", text="STATE_1")
        self.Data_table.heading("ZIP_1", text="ZIP_1")
        self.Data_table.heading("PH_NO1", text="PH_NO1")
        self.Data_table.heading("COUNTRY_1", text="COUNTRY_1")
        self.Data_table.heading("SEX_1", text="SEX_1")
        self.Data_table.heading("D_BIRTH_11", text="D_BIRTH_11")
        self.Data_table.heading("HEIGHT", text="HEIGHT")
        self.Data_table.heading("WEIGHT", text="WEIGHT")
        self.Data_table.heading("BLOOD_GROUP", text="BLOOD_GROUP")
        self.Data_table.heading("BILLING_NAME", text="BILLING_NAME")
        self.Data_table.heading("SHIPPER_NAME", text="SHIPPER_NAME")
        self.Data_table.heading("CITY_2", text="CITY_2")
        self.Data_table.heading("STATE_2", text="STATE_2")
        self.Data_table.heading("ZIP_2", text="ZIP_2")
        self.Data_table.heading("COUNTRY_2", text="COUNTRY_2")
        self.Data_table.heading("PH_NO2", text="PH_NO2")
        self.Data_table.heading("ALCOHOLIC", text="ALCOHOLIC")
        self.Data_table.heading("SMOKER", text="SMOKER")
        self.Data_table.heading("PAST_SURG", text="PAST_SURG")
        self.Data_table.heading("DIABETIC", text="DIABETIC")
        self.Data_table.heading("ALLERGISED", text="ALLERGISED")
        self.Data_table.heading("POLICY_NO", text="POLICY_NO")
        self.Data_table.heading("D_B_LIFE_ASSURE_28", text="D_B_LIFE_ASSURE_28")
        self.Data_table.heading("P_INST", text="P_INST")
        self.Data_table.heading("NAME_P_HOLDER", text="NAME_P_HOLDER")
        self.Data_table.heading("STM_NAME", text="STM_NAME")
        self.Data_table.heading("STM_CODE", text="STM_CODE")
        self.Data_table.heading("DOB_33", text="DOB_33")
        self.Data_table.heading("SEX_2", text="SEX_2")
        self.Data_table.heading("CARD_NAME", text="CARD_NAME")
        self.Data_table.heading("MEDICINE", text="MEDICINE")
        self.Data_table.heading("DOSAGE", text="DOSAGE")
        self.Data_table.heading("TABLETS", text="TABLETS")
        self.Data_table.heading("PILL_RATE", text="PILL_RATE")
        self.Data_table.heading("COST", text="COST")
        self.Data_table.heading("SHIPPING_COST", text="SHIPPING_COST")
        self.Data_table.heading("TOTAL_AMT", text="TOTAL_AMT")
        self.Data_table.heading("REMARK", text="REMARK")

        self.Data_table['show'] = 'headings'
        self.Data_table.column("NO.", anchor=CENTER, width=50)
        self.Data_table.column("RECORD_NO", anchor=CENTER, width=120)
        self.Data_table.column("CUSTOMER_NAME", anchor=W, width=220)
        self.Data_table.column("EMAIL_ADDRESS", anchor=CENTER, width=100)
        self.Data_table.column("RES_ADDRESS", anchor=W, width=320)
        self.Data_table.column("CITY_1", anchor=W, width=150)
        self.Data_table.column("STATE_1", anchor=W, width=140)
        self.Data_table.column("ZIP_1", anchor=CENTER, width=120)
        self.Data_table.column("PH_NO1", anchor=CENTER, width=100)
        self.Data_table.column("COUNTRY_1", anchor=W, width=120)
        self.Data_table.column("SEX_1", anchor=W, width=100)
        self.Data_table.column("D_BIRTH_11", anchor=CENTER, width=140)
        self.Data_table.column("HEIGHT", anchor=CENTER, width=90)
        self.Data_table.column("WEIGHT", anchor=CENTER, width=90)
        self.Data_table.column("BLOOD_GROUP", anchor=CENTER, width=90)
        self.Data_table.column("BILLING_NAME", anchor=W, width=220)
        self.Data_table.column("SHIPPER_NAME", anchor=W, width=100)
        self.Data_table.column("CITY_2", anchor=W, width=150)
        self.Data_table.column("STATE_2", anchor=W, width=140)
        self.Data_table.column("ZIP_2", anchor=CENTER, width=120)
        self.Data_table.column("COUNTRY_2", anchor=W, width=120)
        self.Data_table.column("PH_NO2", anchor=CENTER, width=100)
        self.Data_table.column("ALCOHOLIC", anchor=CENTER, width=100)
        self.Data_table.column("SMOKER", anchor=CENTER, width=100)
        self.Data_table.column("PAST_SURG", anchor=CENTER, width=100)
        self.Data_table.column("DIABETIC", anchor=CENTER, width=100)
        self.Data_table.column("ALLERGISED", anchor=CENTER, width=100)
        self.Data_table.column("POLICY_NO", anchor=W, width=200)
        self.Data_table.column("D_B_LIFE_ASSURE_28", anchor=CENTER, width=140)
        self.Data_table.column("P_INST", width=100)
        self.Data_table.column("NAME_P_HOLDER", anchor=W, width=220)
        self.Data_table.column("STM_NAME", anchor=W, width=150)
        self.Data_table.column("STM_CODE", anchor=W, width=200)
        self.Data_table.column("DOB_33", anchor=CENTER, width=140)
        self.Data_table.column("SEX_2", anchor=W, width=100)
        self.Data_table.column("CARD_NAME", anchor=W, width=120)
        self.Data_table.column("MEDICINE", anchor=W, width=100)
        self.Data_table.column("DOSAGE", anchor=CENTER, width=90)
        self.Data_table.column("TABLETS", anchor=CENTER, width=90)
        self.Data_table.column("PILL_RATE", anchor=CENTER, width=90)
        self.Data_table.column("COST", anchor=CENTER, width=90)
        self.Data_table.column("SHIPPING_COST", anchor=CENTER, width=90)
        self.Data_table.column("TOTAL_AMT", anchor=CENTER, width=90)
        self.Data_table.column("REMARK", anchor=W, width=300)
        self.Data_table.bind("<ButtonRelease-1>", self.get_cursor)
        n = []
        df = openpyxl.load_workbook(r'.\res\Rishu.xlsx')
        read = df.active
        m = read.max_column
        k = read.max_row
        for j in range(2, k + 1):

            for i in range(1, m + 1):
                cell_obj = read.cell(row=j, column=i)
                row = (cell_obj.value)
                n.append(row)

            self.Data_table.insert("", END, values=n)
            n = []

    def get_cursor(self, ev):
        cursor_row = self.Data_table.focus()
        contents = self.Data_table.item(cursor_row)
        self.row = contents['values']
        self.var_Rec_No.set(self.row[1]),
        self.var_cust_name.set(self.row[2]),
        self.var_email_add.set(self.row[3]),
        self.var_res_add.set(self.row[4]),
        self.var_city_1.set(self.row[5]),
        self.var_state_1.set(self.row[6]),
        self.var_zip.set(self.row[7]),
        self.var_ph_no_1.set(self.row[8]),
        self.var_country_1.set(self.row[9]),
        self.var_sex_1.set(self.row[10]),
        self.var_d_birth.set(self.row[11]),
        self.var_height.set(self.row[12]),
        self.var_weidth.set(self.row[13]),
        self.var_Blood_gp.set(self.row[14]),
        self.var_billing_name.set(self.row[15]),
        self.var_shipper_name.set(self.row[16]),
        self.var_city_2.set(self.row[17]),
        self.var_state_2.set(self.row[18]),
        self.var_zip_2.set(self.row[19]),
        self.var_country_2.set(self.row[20]),
        self.var_ph_no_2.set(self.row[21]),
        self.var_alcoholic.set(self.row[22]),
        self.var_smoker.set(self.row[23]),
        self.var_part_sung.set(self.row[24]),
        self.var_diabetic.set(self.row[25]),
        self.var_allergised.set(self.row[26]),
        self.var_policy_no.set(self.row[27]),
        self.var_D_b_lifeassure.set(self.row[28]),
        self.var_p_inst.set(self.row[29], )
        self.var_name_p_holder.set(self.row[30]),
        self.var_stm_name.set(self.row[31]),
        self.var_smt_code.set(self.row[32]),
        self.var_dob.set(self.row[33]),
        self.var_sex_2.set(self.row[34]),
        self.var_crd_name.set(self.row[35]),
        self.var_medicine.set(self.row[36]),
        self.var_Dosage.set(self.row[37]),
        self.var_Tablets.set(self.row[38]),
        self.var_pill_rate.set(self.row[39]),
        self.var_cost.set(self.row[40]),
        self.var_shipimg_cost.set(self.row[41]),
        self.var_Total_amount.set(self.row[42]),
        self.var_Remark.set(self.row[43])
        self.top.destroy()

    def update(self):
        f=('.\res\Rishu.xlsx')
        workbook = openpyxl.load_workbook(f)

        worksheet = workbook.active

        number_of_rows = worksheet.max_row

        number_of_columns = worksheet.max_column

        replacementTextKeyPairs = {str(self.row[1]): self.var_Rec_No.get(), self.row[2]: self.var_cust_name.get(),
                                   self.row[3]: self.var_email_add.get(),self.row[4]:self.var_res_add.get(),self.row[5]:self.var_city_1.get(),self.row[6]:self.var_state_1.get(),
                                   self.row[7]:self.var_zip.get(),self.row[8]:self.var_ph_no_1.get(),self.row[9]:self.var_country_1.get(),self.row[10]:self.var_sex_1.get(),
                                   self.row[11]:self.var_d_birth.get(),self.row[12]:self.var_height.get()

                                   }

        for i in range(number_of_columns):
            for k in range(1):

                cellValue = str(worksheet[get_column_letter(i + 1) + str(k + self.row[0])].value)

                for key in replacementTextKeyPairs.keys():

                    if str(cellValue) == key:
                        newCellValue = replacementTextKeyPairs.get(key)
                        worksheet[get_column_letter(i + 1) + str(k + self.row[0])] = str(newCellValue)

        workbook.save('.\res\Rishu.xlsx')
        self.clear()

    def ext(self):
        msg = messagebox.askquestion("que", "Do you want to exit ?")
        if msg == "yes":
            self.root.destroy()
        else:
            pass

    def rest(self):
        self.__init__(root)

    def show_file(self):
        data = [
            self.File_name.get(),
            self.no.get(),
            self.var_Rec_No.get(),
            self.var_cust_name.get(),
            self.var_email_add.get(),
            self.var_res_add.get(),
            self.var_city_1.get(),
            self.var_state_1.get(),
            self.var_zip.get(),
            self.var_ph_no_1.get(),
            self.var_country_1.get(),
            self.var_sex_1.get(),
            self.var_d_birth.get(),
            self.var_height.get(),
            self.var_weidth.get(),
            self.var_Blood_gp.get(),
            self.var_billing_name.get(),
            self.var_shipper_name.get(),
            self.var_city_2.get(),
            self.var_state_2.get(),
            self.var_zip_2.get(),
            self.var_country_2.get(),
            self.var_ph_no_2.get(),
            self.var_alcoholic.get(),
            self.var_smoker.get(),
            self.var_part_sung.get(),
            self.var_diabetic.get(),
            self.var_allergised.get(),
            self.var_policy_no.get(),
            self.var_D_b_lifeassure.get(),
            self.var_p_inst.get(),
            self.var_name_p_holder.get(),
            self.var_stm_name.get(),
            self.var_smt_code.get(),
            self.var_dob.get(),
            self.var_sex_2.get(),
            self.var_crd_name.get(),
            self.var_medicine.get(),
            self.var_Dosage.get(),
            self.var_Tablets.get(),
            self.var_pill_rate.get(),
            self.var_cost.get(),
            self.var_shipimg_cost.get(),
            self.var_Total_amount.get(),
            self.var_Remark.get()

        ]

        p = []

        workbook_name ='.\res\test.xlsx'
        wb = load_workbook(workbook_name)
        page = wb.active
        m = page.max_row
        for i in range(2, m + 1):
            line = page.cell(column=2, row=i)
            mh = line.value
            p.append(mh)
        # print(p)

        new_companies = [data]
        for info in new_companies:
            page.append(info)

        wb.save(filename=workbook_name)

        p = []

    def box_list(self):
        b = []
        df = openpyxl.load_workbook(r".\res\test.xlsx")
        read = df.active
        m = read.max_row
        for i in range(2, m + 1):
            reads = read.cell(column=1, row=i)
            load_file = (reads.value)
            b.append(load_file)
        # print(b)
        self.res = []
        for i in b:
            if i not in self.res:
                self.res.append(i)

    def serch_itm(self):
        self.Data_table.delete(*self.Data_table.get_children())
        b = []
        search_words = self.box_item.get()

        wb = load_workbook(r".\res\test.xlsx")
        ws = wb.active

        for row in ws.rows:
            if row[0].value == search_words:
                for cell in row:
                    vl = cell.value
                    b.append(vl)
                self.Data_table.insert("", END, values=b[1:])
                b = []


root=tk.Tk()
ob=Masala(root)
mainloop()